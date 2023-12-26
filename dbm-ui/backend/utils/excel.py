# -*- coding: utf-8 -*-
"""
TencentBlueKing is pleased to support the open source community by making 蓝鲸智云-DB管理系统(BlueKing-BK-DBM) available.
Copyright (C) 2017-2023 THL A29 Limited, a Tencent company. All rights reserved.
Licensed under the MIT License (the "License"); you may not use this file except in compliance with the License.
You may obtain a copy of the License at https://opensource.org/licenses/MIT
Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on
an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the
specific language governing permissions and limitations under the License.
"""

from collections import defaultdict
from io import BytesIO
from typing import Any, Dict, List, Union

import openpyxl
from django.http.response import HttpResponse, StreamingHttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.writer.excel import save_virtual_workbook


class ExcelHandler:
    """
    封装常用的excel处理函数
    """

    @classmethod
    def _adapt_sheet_weight_height(cls, sheet: Worksheet, first_header_row: int = 1):
        """
        自动调整excel的行高度和列宽度
        :param sheet: worksheet
        :param first_header_row: 表头开始的行数(有可能前几行是annotations)
        """

        # 列宽计算公式: (max_str_len).encode("gbk") * 1.3 (同一单元格中通过\n分割的字符视为独立的长度)
        # 注：这里为什么用gbk编码呢？因为在gbk中一个中文字符为两个字节，英文为1个字节，而在excel中一个中文字符的大小接近于两个英文小写字符
        # 因此用gbk编码来统计字符大小作为字符的宽度。
        # 事实上大写英文字符，小数点，特殊符号大小也各不相同，但是openpyxl中没有自适应宽度的函数，为了不便于问题复杂化，就先这样简约处理吧。
        row_num, col_num = sheet.max_row, sheet.max_column
        max_col_dimensions: List[int] = [0 for _ in range(col_num + 1)]
        for row in range(first_header_row, row_num + 1):
            for col in range(1, col_num + 1):
                # 自动调整行高度
                cell = sheet.cell(row, col)
                cell.alignment = Alignment(wrapText=True)

                cell_str_len_list = [len(cell_str.encode("gbk")) for cell_str in str(cell.value).split("\n")]
                max_col_dimensions[col] = max(max_col_dimensions[col], max(cell_str_len_list) * 1.3)

        # 自动调整列宽度
        for col in range(1, col_num + 1):
            sheet.column_dimensions[chr(ord("A") + col - 1)].width = max_col_dimensions[col]

    @classmethod
    def paser(cls, excel: BytesIO, header_row: int = 0, sheet_name: str = "") -> List[Dict]:
        """
        - 解析excel文件为数据字典
        :param excel: excel二进制文件
        :param header_row: excel头部的行数(有可能存在excel的前几行是annotations的情况，因此需要用户指定)
        :param sheet_name: sheet的名称
        """

        if not sheet_name:
            excel_rows = list(openpyxl.load_workbook(excel).active.rows)
        else:
            excel_rows = list(openpyxl.load_workbook(excel)[sheet_name].rows)

        header_list = [header.value for header in excel_rows[header_row]]
        excel_data_dict__list = []
        for content_row in excel_rows[header_row + 1 :]:
            content_list = [str(content.value) for content in content_row]
            excel_data_dict__list.append(dict(zip(header_list, content_list)))

        return excel_data_dict__list

    @classmethod
    def paser_matrix(cls, excel: Union[BytesIO, str]) -> Dict[str, Dict[str, Any]]:
        """
        - 解析excel文件为二维矩阵，默认第一行和第一列是索引指标
        :param excel: excel二进制文件
        """

        matrix_data: Dict[str, Dict[str, Any]] = defaultdict(dict)
        wb = openpyxl.load_workbook(excel)
        for sheet in wb.worksheets:
            for row in range(2, sheet.max_row + 1):
                for col in range(2, sheet.max_column + 1):
                    if sheet.cell(row, col).value:
                        matrix_data[sheet.cell(1, col).value][sheet.cell(row, 1).value] = sheet.cell(row, col).value

        return matrix_data

    @classmethod
    def serialize(
        cls,
        data_dict__list: List[Dict],
        template: str = None,
        headers: List = None,
        header_style: List = None,
        match_header: bool = False,
    ) -> Workbook:
        """
        - 将数据字典序列化为excel对象
        :param data_dict__list: 数据字典
        :param template: excel模板路径(优先以模板的头部样式作为excel的头部)
        :param headers: excel数据头 [{"id": "header_id", "name": "header_name"}]
        :param header_style: excel的头部样式(颜色)
        :param match_header: 数据是否匹配表头，如果为True，则根据 header 严格匹配列名，若不存在，则在该 cell 填充空
        """

        wb: Workbook = Workbook()
        sheet: Worksheet = wb.active
        first_data_row: int = 2

        if template:
            # 导入template，设置excel头部和样式(考虑模板的第一行为注释，TODO：可读性较差，这部分需改进)
            first_data_row = 3
            wb = openpyxl.load_workbook(template)
            sheet = wb.active
        elif headers:
            # 如果没有template，则根据给定的header和header颜色来设置头部和样式
            for col, header in enumerate(headers):
                header = header if isinstance(header, str) else header["name"]
                cell = sheet.cell(1, col + 1, str(header))
                if header_style:
                    cell.fill = PatternFill("solid", fgColor=header_style[header])

        # 数据写入单元格
        for row, data_dict in enumerate(data_dict__list):
            if match_header:
                # 如果match_header为True，则根据 header 严格匹配列名，若不存在，则在该 cell 填充空
                for col, header in enumerate(headers):
                    header_id = header if isinstance(header, str) else header["id"]
                    if header_id not in data_dict:
                        sheet.cell(row + first_data_row, col + 1).value = ""
                    else:
                        sheet.cell(row + first_data_row, col + 1, str(data_dict[header_id]))
            else:
                for col, value in enumerate(list(data_dict.values())):
                    sheet.cell(row + first_data_row, col + 1, str(value))

        # 自适应设置行高和列宽
        cls._adapt_sheet_weight_height(sheet=sheet, first_header_row=first_data_row - 1)

        return wb

    @classmethod
    def response(cls, wb: Workbook, excel_name: str) -> HttpResponse:
        """
        - 返回excel文件的HttpResponse
        :param wb: excel的Workbook
        :param excel_name: excel文件名
        """

        # 设置response格式并写入bytes
        response = HttpResponse(
            content=save_virtual_workbook(wb),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment;filename={excel_name}"
        return response

    @classmethod
    def stream_response(cls, wb: Workbook, excel_name: str) -> StreamingHttpResponse:
        # TODO 如果excel文件过大，需采用流式返回，待实现
        raise NotImplementedError
